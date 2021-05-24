#def send_email():
import openpyxl

def read_excel(File):
    workbook = openpyxl.load_workbook(filename="f{File}")
    sheet = workbook.active
    int max_row = sheet.max_row
    for i in range(0..max_row):
        



if __name__ == "__main__":
    global Excel_file 
    Excel_file = input("Enter path of Excel File:\n")
    read_excel(Excel_file)

    
