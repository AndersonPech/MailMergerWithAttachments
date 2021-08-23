import openpyxl
import merge_email
import data
import setData

def read_excel(File):
    '''
    Reads the excel file by row. 
    Precondition: Excel file is in accurate format
    Postcondition: Email sent
    '''
    try:
        workbook = openpyxl.load_workbook(filename=File)
        sheet = workbook.active
        data.Excel_file = sheet
    except FileNotFoundError:
        print("File not found")

    ##print(data.Excel_file.cell(row=2,column=2).value)
    max_row = sheet.max_row + 1
    for i in range(2, max_row):
        setData.setMessage(
        f"Dear {str(data.Excel_file.cell(row=i, column=1).value)} ,\n" 
        f"Please see the attached invoice\n"
        )
        setData.setSubject("TestEmail")
        merge_email.send_email(data.Excel_file.cell(row=i, column=2).value, 
            data.Excel_file.cell(row=i, column=3).value)